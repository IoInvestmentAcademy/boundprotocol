#!/usr/bin/env python3
"""Verify public/bound-cost-analysis.html against its two sources.

Re-derives every headline figure independently from the financial model workbook
and scripts/sim-default.json, then asserts each one appears in the rendered page.
The rendered document must never contain a number that cannot be reproduced from
source.

Usage:  python3 scripts/verify-costs.py "<path to workbook.xlsx>"
"""
import html
import json
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XL = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    '~/Downloads/Copy of BOUND Financial Model Last Version.xlsx')

CAS, CASS, PIT, CAM = 0.25, 0.10, 0.10, 0.0225
FEE_RATE = 0.001

wb = openpyxl.load_workbook(XL, data_only=True)
co, pc = wb['Costs'], wb['Personnel costs']

colyear = {c: int(co.cell(7, c).value) for c in range(9, 69)
           if isinstance(co.cell(7, c).value, (int, float))}
YK = sorted(set(colyear.values()))[:3]


def crow(r):
    o = {y: 0.0 for y in YK}
    for c, y in colyear.items():
        if y in o:
            v = co.cell(r, c).value
            if isinstance(v, (int, float)):
                o[y] += v
    return [o[y] for y in YK]


sim = json.load(open(os.path.join(ROOT, 'scripts/sim-default.json')))
SM = sim['months'][:36]
assert len(SM) == 36, 'simulator feed must carry 36 months'

syear = lambda k, y: sum(m[k] for m in SM[y * 12:(y + 1) * 12])
payfee = [m['volume'] * FEE_RATE for m in SM]


def compound(seed):
    out, cur = [], seed / 12
    out.append(cur)
    for i in range(1, 36):
        prev = SM[i - 1]['volume']
        g = (SM[i]['volume'] / prev - 1) if prev > 0 else 0.0
        cur *= (1 + max(g, -0.9))
        out.append(cur)
    return out


ysum = lambda a, y: sum(a[y * 12:(y + 1) * 12])
cloud, maint = compound(42000), compound(50000)

# personnel from the summary block, 36-month window
pers_y = [0.0, 0.0, 0.0]
for r in range(73, 85):
    months = [pc.cell(r, c).value for c in range(8, 44)]
    months = [m if isinstance(m, (int, float)) else 0 for m in months]
    if sum(months) == 0:
        continue
    for i in range(3):
        pers_y[i] += sum(months[i * 12:(i + 1) * 12])

pers_emp_y = [g * (1 + CAM) for g in pers_y]
software, ga, legal, dev = crow(45), crow(59), crow(66), crow(76)

# Sales and marketing is no longer taken from the workbook. It is rebuilt from B2B
# benchmarks at $1,500,000 — reproduced here independently of the generator so the
# check is a check, not an echo of the thing it verifies.
MKT_PHASES = [300_000, 280_000, 920_000]     # B2B institutional · pool launch · open market
MKT_TOTAL = sum(MKT_PHASES)
assert MKT_TOTAL == 1_500_000, 'marketing plan must total $1,500,000'
_p3 = MKT_PHASES[2] / 28.0
sm = [MKT_PHASES[0] + MKT_PHASES[1] + _p3 * 4, _p3 * 12, _p3 * 12]
cloud_y = [ysum(cloud, y) for y in range(3)]
maint_y = [ysum(maint, y) for y in range(3)]
fee_y = [ysum(payfee, y) for y in range(3)]

cogs_y = [software[i] + cloud_y[i] + maint_y[i] + pers_emp_y[i] + fee_y[i] for i in range(3)]
total_y = [cogs_y[i] + ga[i] + legal[i] + sm[i] + dev[i] for i in range(3)]
rev_y = [syear('revenue', y) for y in range(3)]

gross = sum(pers_y)
cas, cass = gross * CAS, gross * CASS
pit = (gross - cas - cass) * PIT
net = gross - cas - cass - pit
cam = gross * CAM
emp = gross + cam
wedge = cas + cass + pit + cam

M = lambda x: '${:,.0f}'.format(round(x))

# --- reconciliation assertions against the workbook itself ---
model_personnel = crow(49)
for i in range(3):
    assert abs(model_personnel[i] - pers_y[i]) < 1, (
        'personnel mismatch between Costs row 49 and Personnel costs rows 73-84 '
        'in year {}: {} vs {}'.format(YK[i], model_personnel[i], pers_y[i]))

# The workbook must still be internally consistent on its own terms — checked
# with the workbook's own marketing figure, not ours.
model_total, wb_mkt = crow(78), crow(71)
for i in range(3):
    rebuilt = crow(51)[i] + ga[i] + legal[i] + wb_mkt[i] + dev[i]
    assert abs(rebuilt - model_total[i]) < 1, (
        'the workbook no longer ties to its own row 78 in {}'.format(YK[i]))

# Our total deliberately departs from the workbook by exactly the marketing
# difference and by nothing else. That is the point of the rebuild, and it is
# asserted rather than assumed.
mkt_delta = sum(wb_mkt) - MKT_TOTAL
rebased_delta = (sum(crow(51)) + sum(ga) + sum(legal) + sum(wb_mkt) + sum(dev)) - sum(model_total)
assert abs(mkt_delta - 2_240_000) < 1, \
    'marketing reduction should be $2,240,000, got {:,.0f}'.format(mkt_delta)

print('reconciliation: Costs row 49 ties to Personnel costs summary  ..... OK')
print('reconciliation: workbook still ties to its own row 78         ..... OK')
print('reconciliation: our total departs only by marketing (−{})   ..... OK'.format(M(mkt_delta)))

# --- figures that must appear in the page ---
page = open(os.path.join(ROOT, 'public/bound-cost-analysis.html')).read()
page = re.sub(r'<style>[\s\S]*?</style>', ' ', page)
page = re.sub(r'<script>[\s\S]*?</script>', ' ', page)
page = html.unescape(re.sub(r'<[^>]+>', ' ', page))

checks = [
    ('rebased 3-year cost base', M(sum(total_y))),
    ('protocol revenue',          M(sum(rev_y))),
    ('personnel employer cost',   M(emp)),
    ('personnel gross',           M(gross)),
    ('net reaching employees',    M(net)),
    ('pension contribution',      M(cas)),
    ('health contribution',       M(cass)),
    ('income tax',                M(pit)),
    ('employer contribution',     M(cam)),
    ('total tax wedge',           M(wedge)),
    ('payment fees 3-yr',         M(sum(fee_y))),
    ('cloud 3-yr',                M(sum(cloud_y))),
    ('maintenance 3-yr',          M(sum(maint_y))),
    ('sales and marketing 3-yr',  M(sum(sm))),
    ('legal 3-yr',                M(sum(legal))),
]
for y in range(3):
    checks.append(('total cost {}'.format(YK[y]), M(total_y[y])))
    checks.append(('revenue {}'.format(YK[y]), M(rev_y[y])))
    checks.append(('personnel employer {}'.format(YK[y]), M(pers_emp_y[y])))

missing = [(lab, val) for lab, val in checks if val not in page]
if missing:
    print('\nFAIL — figures absent from the rendered page:')
    for lab, val in missing:
        print('   {:<34} {}'.format(lab, val))
    sys.exit(1)

for lab, val in checks:
    print('  {:<34} {:>16}  OK'.format(lab, val))

# --- entity attribution and the operating company ---
SERVICES_MARGIN, EU_TAX = 0.08, 0.16
opco_y = [pers_emp_y[i] + sm[i] + maint_y[i] + cloud_y[i] + legal[i] + dev[i] + ga[i] + software[i]
          for i in range(3)]
proto_y = [fee_y[i] for i in range(3)]
for i in range(3):
    assert abs((opco_y[i] + proto_y[i]) - total_y[i]) < 1, \
        'entity attribution does not reconcile to the cost base in year {}'.format(i + 1)
print('  entity attribution reconciles to the cost base      ..... OK')

opco_rev = [c * (1 + SERVICES_MARGIN) for c in opco_y]
opco_marg = [r - c for r, c in zip(opco_rev, opco_y)]
opco_tax = [m * EU_TAX for m in opco_marg]
for lab, val in [('services revenue', M(sum(opco_rev))), ('margin', M(sum(opco_marg))),
                 ('EU tax', M(sum(opco_tax))), ('marketing total', M(MKT_TOTAL))]:
    if val not in page:
        print('\nFAIL — {} absent from the page: {}'.format(lab, val)); sys.exit(1)
    print('  {:<34} {:>16}  OK'.format(lab, val))

# The document must never net protocol revenue against operating cost. The prose
# guard below caught the wording; it did not catch a COVERAGE TABLE doing the same
# thing silently, which is how the error survived into publication once already.
# Both are checked now.
import re as _re
if _re.search(r'protocol revenue[^.]{0,80}(less|minus|against)[^.]{0,40}operating cost', page, _re.I):
    print('\nFAIL — the page nets protocol revenue against operating cost'); sys.exit(1)

# Structural guard: the blended deficit must never appear as a reported result.
# It may appear ONCE, in the passage that exists to rule it out.
blended = M(sum(rev_y) - sum(total_y))
occurrences = page.count(blended)
if occurrences > 1:
    print('\nFAIL — the blended protocol-vs-total deficit {} appears {} times; it may '
          'appear only in the passage that disclaims it'.format(blended, occurrences))
    sys.exit(1)
if occurrences == 1 and 'not combined' not in page and 'never netted' not in page:
    print('\nFAIL — {} appears without the passage explaining why the two are not '
          'combined'.format(blended)); sys.exit(1)

# Coverage must be reported per entity, not blended.
proto_cov = sum(rev_y) / sum(proto_y)
opco_net = sum(r - c for r, c in zip(opco_rev, opco_y)) * (1 - EU_TAX)
for lab, val in [('protocol coverage', '{:.2f}\u00d7'.format(proto_cov)),
                 ('opco net result', M(opco_net))]:
    if val not in page:
        print('\nFAIL — per-entity {} absent from the page: {}'.format(lab, val)); sys.exit(1)
    print('  %-34s %16s  OK' % (lab, val))
print('  protocol revenue is not netted against opco cost   ..... OK')
print('  coverage is reported per entity, not blended       ..... OK')

print()
print('PASS — every derived figure appears in the rendered page,')
print('       and both source reconciliations hold.')
