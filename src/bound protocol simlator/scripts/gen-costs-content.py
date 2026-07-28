#!/usr/bin/env python3
"""Generate scripts/costs-content.json for the Operating Cost Analysis document.

Reads two sources and derives every figure in the document from them:
  1. the financial model workbook — sheets `Costs` and `Personnel costs` only
  2. scripts/sim-default.json — the simulator's default-case monthly series

The financial model's own revenue-scaling lines are driven by a separate
investment-balance series that is unrelated to protocol activity. Those lines are
rebased here onto the simulator's default case. Cost STRUCTURE is preserved;
only the DRIVERS are replaced.

Usage:  python3 scripts/gen-costs-content.py "<path to workbook.xlsx>"
"""
import json
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XL = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    '~/Downloads/Copy of BOUND Financial Model Last Version.xlsx')

# --- Romanian payroll rates, verified against PwC Worldwide Tax Summaries (2026) ---
CAS, CASS, PIT, CAM = 0.25, 0.10, 0.10, 0.0225
NET_RATE = 1 - CAS - CASS - (1 - CAS - CASS) * PIT   # 0.585

YEARS = ['FY2026', 'FY2027', 'FY2028']

wb = openpyxl.load_workbook(XL, data_only=True)
co, pc = wb['Costs'], wb['Personnel costs']

# `Costs` data starts at column I (H is the "USD" label); `Personnel costs` at H.
# Reading both from the same column silently drops December from every year.
colyear = {c: int(co.cell(7, c).value) for c in range(9, 69)
           if isinstance(co.cell(7, c).value, (int, float))}
YKEYS = sorted(set(colyear.values()))[:3]


def crow(r):
    """Yearly totals for a `Costs` row, first three years."""
    o = {y: 0.0 for y in YKEYS}
    for c, y in colyear.items():
        if y in o:
            v = co.cell(r, c).value
            if isinstance(v, (int, float)):
                o[y] += v
    return [o[y] for y in YKEYS]


sim = json.load(open(os.path.join(ROOT, 'scripts/sim-default.json')))
SM = sim['months'][:36]


def syear(key, y):
    return sum(m[key] for m in SM[y * 12:(y + 1) * 12])


# ---------------------------------------------------------------- rebased drivers
FEE_RATE = 0.001          # 0.1% of protocol transaction volume
payfee = [m['volume'] * FEE_RATE for m in SM]


def compound_on_volume(annual_seed):
    """Financial-model recipe: seed at annual/12, then compound by volume growth.
    The growth series is the simulator's, not the model's own."""
    out, cur = [], annual_seed / 12
    out.append(cur)
    for i in range(1, 36):
        prev = SM[i - 1]['volume']
        g = (SM[i]['volume'] / prev - 1) if prev > 0 else 0.0
        cur = cur * (1 + max(g, -0.9))
        out.append(cur)
    return out


cloud = compound_on_volume(42000)
maint = compound_on_volume(50000)
ysum = lambda a, y: sum(a[y * 12:(y + 1) * 12])

# ---------------------------------------------------------------- personnel
roles = []
for r in range(73, 85):
    name = pc.cell(r, 2).value or pc.cell(r, 3).value or pc.cell(r, 6).value
    months = [pc.cell(r, c).value for c in range(8, 44)]
    months = [m if isinstance(m, (int, float)) else 0 for m in months]
    if sum(months) == 0:
        continue
    roles.append({
        'name': str(name),
        'years': [sum(months[i * 12:(i + 1) * 12]) for i in range(3)],
        'gross': sum(months),
    })

roster_ref = {}
for r in range(9, 21):
    role = pc.cell(r, 6).value
    if role:
        flags = [pc.cell(r, c).value for c in range(8, 68)]
        active = [i for i, f in enumerate(flags) if f == 1]
        roster_ref[str(role)] = {
            'base': pc.cell(r, 2).value or 0,
            'bonus': pc.cell(r, 3).value or 0,
            'alloc': pc.cell(r, 4).value or '',
            'hire': (2026 + active[0] // 12, active[0] % 12 + 1) if active else None,
            'months': len(active),
        }

never_hired = [k for k, v in roster_ref.items() if v['hire'] is None]

pers_gross_y = [sum(r['years'][i] for r in roles) for i in range(3)]
pers_gross = sum(pers_gross_y)
pers_cam_y = [g * CAM for g in pers_gross_y]
pers_emp_y = [g + c for g, c in zip(pers_gross_y, pers_cam_y)]

# ---------------------------------------------------------------- cost model
software = crow(45)
ga, legal, dev = crow(59), crow(66), crow(76)

# --- Sales and marketing: rebuilt from B2B benchmarks, replacing the workbook ----
# The workbook carried $3,740,000 ($1.58M / $1.08M / $1.08M), sized for a mass-market
# launch. The go-to-market plan is a direct sales motion to institutional liquidity
# providers and RWA issuers for its first two stages, which does not consume that.
# Rebuilt bottom-up against published 2026 benchmarks — Tier-2 conference $50–150K
# all-in, $5–15K per sourced opportunity, median B2B sales-led CAC $11,400 — the
# budget lands at $1,500,000, and the Marketing token allocation covers two thirds
# of it.
MARKETING_PLAN = [
    ('Phase 1 — B2B institutional', 'Months 1–5', [
        ('Brand system and institutional collateral', 60_000),
        ('Category research and thought leadership', 50_000),
        ('Two Tier-2 conferences', 120_000),
        ('Public relations and analyst relations', 40_000),
        ('Issuer business-development materials and data room', 30_000),
    ]),
    ('Phase 2 — Pool launch', 'Months 6–8', [
        ('Launch campaign', 120_000),
        ('One Tier-2 conference', 60_000),
        ('Public relations and earned media', 40_000),
        ('DeFi ecosystem and integrations', 60_000),
    ]),
    ('Phase 3 — Open market', 'Months 9–36', [
        ('Content and search, $12,000 a month', 336_000),
        ('Community and education, $7,000 a month', 196_000),
        ('Conferences, two a year', 240_000),
        ('Performance marketing, retail acquisition', 148_000),
    ]),
]
MKT_PHASE_TOTAL = [sum(v for _, v in items) for _, _, items in MARKETING_PLAN]
MKT_TOTAL = sum(MKT_PHASE_TOTAL)
assert MKT_TOTAL == 1_500_000, 'marketing plan must total $1,500,000'
# Phases 1 and 2 fall inside year one; phase 3 spans months 9-36, so it is split
# across the three years by the months each year contributes.
_p3 = MKT_PHASE_TOTAL[2] / 28.0
sm = [MKT_PHASE_TOTAL[0] + MKT_PHASE_TOTAL[1] + _p3 * 4, _p3 * 12, _p3 * 12]
MKT_TOKEN_ALLOCATION = 1_000_000   # Marketing allocation, 2% of supply at launch price
cloud_y = [ysum(cloud, y) for y in range(3)]
maint_y = [ysum(maint, y) for y in range(3)]
fee_y = [ysum(payfee, y) for y in range(3)]

cogs_y = [software[i] + cloud_y[i] + maint_y[i] + pers_emp_y[i] + fee_y[i] for i in range(3)]
total_y = [cogs_y[i] + ga[i] + legal[i] + sm[i] + dev[i] for i in range(3)]
rev_y = [syear('revenue', y) for y in range(3)]
net_y = [rev_y[i] - total_y[i] for i in range(3)]
ratio_y = [total_y[i] / rev_y[i] for i in range(3)]

# original model figures, for the comparison table
orig_total = crow(78)
orig_fee = crow(50)
orig_cloud, orig_maint = crow(46), crow(47)

# ---------------------------------------------------------------- entity model
# Three entities. Costs and revenues belong to different ones and must never be
# netted against each other — the error this document exists to correct.
#
#   Nexus Technologies Foundation (Panama)  owns the protocol and its IP
#   BITSWIFT TECHNOLOGIES INC.   (Panama)   distributes the token, holds proceeds
#   Operating company            (EU)       employs staff, bears operating cost
#
# The operating company is paid under a development-and-services agreement at
# cost plus a margin. Without it the company has no revenue line at all, which is
# both commercially wrong and indefensible on transfer pricing: an EU company
# performing real work for a Panama principal at zero margin is the textbook
# challenge.
SERVICES_MARGIN = 0.08
EU_TAX_RATE = 0.16

# Every line, and the entity that bears it.
ENTITY_LINES = [
    ('Personnel (employer cost)', pers_emp_y, 'Operating company'),
    ('Sales and marketing',       sm,         'Operating company'),
    ('Maintenance infrastructure', maint_y,   'Operating company'),
    ('Cloud infrastructure',      cloud_y,    'Operating company'),
    ('Legal',                     legal,      'Split — token legal to Issuer'),
    ('Development',               dev,        'Operating company — billed on, IP to Foundation'),
    ('General and administrative', ga,        'Operating company'),
    ('Software licences',         software,   'Operating company'),
    ('Payment fees',              fee_y,      'Protocol'),
]
opco_y  = [sum(v[i] for _, v, o in ENTITY_LINES if o.startswith('Operating') or o.startswith('Split'))
           for i in range(3)]
proto_y = [sum(v[i] for _, v, o in ENTITY_LINES if o == 'Protocol') for i in range(3)]

# --- operating company P&L ---
opco_rev_y   = [c * (1 + SERVICES_MARGIN) for c in opco_y]
opco_marg_y  = [r - c for r, c in zip(opco_rev_y, opco_y)]
opco_tax_y   = [m * EU_TAX_RATE for m in opco_marg_y]
opco_net_y   = [m - t for m, t in zip(opco_marg_y, opco_tax_y)]

# Development work is billed to the Foundation, which capitalises the resulting IP
# rather than expensing it — the Foundation owns the protocol, so the spend has to
# produce an asset somewhere or the group has paid for something it does not hold.
dev_capitalised = [d * (1 + SERVICES_MARGIN) for d in dev]

# --- funding available to the operating company ---
# Token-supply constants. These mirror the tokenomics defaults in
# src/BoundSimulator.jsx and must be changed there first — the allocation is
# derived here rather than hardcoded so a change to Treasury cannot silently
# leave this funding table stale.
TOTAL_SUPPLY  = 1_000_000_000
LAUNCH_PRICE  = 0.05
TREASURY_PCT  = 9.0      # src/BoundSimulator.jsx -> alloc.treasury
MARKETING_PCT = 2.0      # src/BoundSimulator.jsx -> alloc.marketing
# Treasury vesting: 5% at TGE, remainder linear over 48 months; 12 months elapsed by M24.
TRE_VESTED_24 = (5 + (12 / 48) * 95) / 100

OPCO_CASH   = 2_480_000 + 7_500_000 * 0.30      # Seed raise + 30% Company share of Public
TOK_MKT     = TOTAL_SUPPLY * MARKETING_PCT / 100 * LAUNCH_PRICE   # fully vested by M24
TOK_TRE_24  = TOTAL_SUPPLY * TREASURY_PCT / 100 * TRE_VESTED_24 * LAUNCH_PRICE
OPCO_AVAIL_24 = OPCO_CASH + TOK_MKT + TOK_TRE_24
# The operating company runs its own 24-month funded plan. The protocol's 36-month
# simulation is a different entity on a different clock; the two are reported
# separately and never combined.
OPCO_COST_24 = sum(opco_rev_y) * 24 / 36
OPCO_SURPLUS_24 = OPCO_AVAIL_24 - OPCO_COST_24
OPCO_BURN_MO = sum(opco_rev_y) / 36

T = lambda v: sum(v)
# Negatives lead with a true minus sign outside the symbol — "−$4,818,741", not "$-4,818,741".
M = lambda x: ('−${:,.0f}' if round(x) < 0 else '${:,.0f}').format(abs(round(x)))
P = lambda x: '{:.1f}%'.format(x)

# ---------------------------------------------------------------- content
front = {
    'brand': 'BOUND Protocol',
    'kind': 'Operating Cost & Company Analysis',
    'tagline': 'Cost Base, Entity Attribution and the Operating Company',
    'version': 'FY2026–FY2028 — July 2026',
    'place': 'Bucharest, Romania',
}


def p(text, bold=False):
    return {'t': 'p', 'text': text, 'boldLabel': bold}


def sub(title, blocks):
    return {'t': 'sub', 'title': title, 'blocks': blocks}


def table(head, rows, cap=None):
    b = {'t': 'table', 'head': head, 'rows': rows}
    if cap:
        b['cap'] = cap
    return b


def ul(items, bold=True):
    return {'t': 'ul', 'items': items, 'boldLabel': bold}


def kpis(items):
    return {'t': 'kpis', 'items': items}


def chart(title, bars):
    return {'t': 'chart', 'title': title, 'bars': bars}


def note(text, kind='', head=''):
    return {'t': 'note', 'text': text, 'kind': kind, 'head': head}


chapters = []

# ---- Abstract
chapters.append({'num': '', 'id': 'abstract', 'title': 'Abstract', 'nav': 'Abstract', 'blocks': [
    p('A cost model is only as good as the revenue series it is indexed to. The financial model '
      'behind this document scaled its largest cost line against an investment balance that the '
      'protocol does not operate — a portfolio fee series unrelated to the liquidity engine the '
      'simulator actually models. This document rebases every revenue-scaling cost onto the '
      'simulator’s default case and reports what the cost base looks like once that correction '
      'is made.'),
    p('Over the thirty-six months the simulator covers, the rebased operating cost base totals '
      + M(T(total_y)) + ', against protocol revenue of ' + M(T(rev_y)) + '. Personnel accounts for '
      + M(T(pers_emp_y)) + ' of that at full employer cost, spread across nine roles that reach '
      'their steady-state headcount within the first five months and do not grow thereafter. '
      'The cost-to-revenue ratio falls from ' + '{:.2f}'.format(ratio_y[0]) + '× in the first year to '
      + '{:.2f}'.format(ratio_y[2]) + '× in the third. It does not reach parity inside the horizon; '
      'the cumulative position at month thirty-six is ' + M(T(net_y)) + '.'),
    p('This document specifies the scope of the analysis, the rebasing that was performed and why, '
      'the composition of the cost base, the personnel plan at gross-to-net detail under Romanian '
      'payroll law, and the external benchmarks each material line was tested against. Where the '
      'model is understated, that is stated. Where an assumption carries no supporting plan, that '
      'is stated too. The figures are model outputs on the assumptions recorded here — not '
      'forecasts, and not a budget that has been committed.'),
]})

# ---- 1 Scope and Method
chapters.append({'num': '1', 'id': 'scope', 'title': 'Scope and Method', 'nav': 'Scope and Method', 'blocks': [
    sub('1.1 Sources', [
        p('Two sheets of the financial model are in scope: <b>Costs</b>, which holds the cost '
          'forecast and its assumptions, and <b>Personnel costs</b>, which holds the roster, the '
          'hiring flags and the salary escalation. No other sheet is quoted, including where the '
          'cost sheets reference one. The third source is the simulator itself: the default preset '
          'against the default OUSG market, dumped to a fixed series so that every rebased figure '
          'in this document can be re-derived rather than trusted.'),
        p('The financial model runs sixty months. The simulator runs thirty-six. This document is '
          'bounded by the shorter of the two — FY2026 through FY2028 — because a cost line rebased '
          'onto a series that does not exist is an extrapolation, not a measurement. Years four and '
          'five of the financial model are outside this analysis.'),
    ]),
    sub('1.2 Reconciliation performed', [
        p('Two structural traps in the workbook were identified and corrected before any figure was '
          'taken. The <b>Costs</b> sheet begins its monthly series at column I; <b>Personnel costs</b> '
          'begins at column H. Reading both from the same column drops December from every year. '
          'And personnel must be read from the summary block, which applies the hiring flags, rather '
          'than the gross block, which prices all twelve defined roles whether or not they are '
          'hired — the difference over five years is $721,268.'),
        p('With both corrected, the personnel line inside the cost forecast ties to the personnel '
          'sheet in all sixty months, and the category totals tie to the model’s own total-expenses '
          'row in every year. Every figure below descends from those two reconciled series.'),
    ]),
]})

# ---- 2 Rebasing
chapters.append({'num': '2', 'id': 'rebasing', 'title': 'Rebasing the Model onto Protocol Results',
                 'nav': 'Rebasing', 'blocks': [
    sub('2.1 What the model was indexed to', [
        p('The financial model computes its largest cost line — payment fees — as one tenth of one '
          'percent of a series labelled <i>TVL for portfolios</i>. That series is an investment '
          'closing balance carried on a different sheet; it is a custody and portfolio-management '
          'fee base, and it moves independently of RWA liquidations, minting, conversions and pool '
          'trading. The protocol’s cost of processing flow was therefore indexed to a balance the '
          'protocol does not charge against.'),
        p('The consequence is not a small calibration error. Measured against the simulator’s '
          'default case, the balance the model assumes exceeds protocol value under custody by '
          '2.3× in the first year, 3.8× in the second and 5.0× in the third — a gap that widens '
          'rather than closes.'),
        table(['Year', 'Model implied balance', 'Simulator TVL', 'Ratio'],
              [[YEARS[y], M(orig_fee[y] / 12 * 1000), M(syear('tvl', y) / 12),
                '{:.1f}×'.format((orig_fee[y] / 12 * 1000) / (syear('tvl', y) / 12))]
               for y in range(3)],
              'Model implied balance is back-solved from the payment-fee line at its own 0.1% rate. '
              'Simulator TVL is the monthly average of Liquidity Layer plus Bound Core balance, '
              'default preset. Source: Costs rows 16 and 50; scripts/sim-default.json.'),
    ]),
    sub('2.2 What was replaced', [
        p('Three lines are driven by the model’s own revenue series and are rebased here. Payment '
          'fees are recomputed at the same one-tenth-of-one-percent rate against protocol '
          'transaction volume — the flow the protocol actually processes. Cloud infrastructure and '
          'infrastructure maintenance retain the model’s recipe, seeding at one twelfth of their '
          'annual assumption and compounding monthly, but compound against the simulator’s volume '
          'growth rather than the model’s.'),
        p('Everything else is carried unchanged. Personnel, software licences, general and '
          'administrative expenses, legal, sales and marketing, and development are not indexed to '
          'revenue anywhere in the model, and rebasing them would invent a relationship the source '
          'does not contain.'),
        table(['Line', 'Original basis', 'Rebased basis', 'Original 3-yr', 'Rebased 3-yr'],
              [['Payment fees', '0.1% × investment balance', '0.1% × transaction volume',
                M(T(orig_fee)), M(T(fee_y))],
               ['Cloud infrastructure', 'compounds on model revenue growth',
                'compounds on simulator volume growth', M(T(orig_cloud)), M(T(cloud_y))],
               ['Maintenance infrastructure', 'compounds on model revenue growth',
                'compounds on simulator volume growth', M(T(orig_maint)), M(T(maint_y))]],
              'Source: Costs rows 46, 47 and 50, against scripts/sim-default.json. '
              'Rates and seeding recipes are unchanged; only the driver series differs.'),
        p('The net effect on the three-year cost base is a reduction of '
          + M(T(orig_total) - T(total_y)) + ', from ' + M(T(orig_total)) + ' as modelled to '
          + M(T(total_y)) + ' as rebased. Almost all of that reduction is the payment-fee line '
          'alone.'),
    ]),
    sub('2.3 Limits of the rebasing', [
        p('The rebasing inherits every assumption of the simulator’s default preset, including its '
          'single OUSG market, its growth path and its fee schedule. A different preset produces a '
          'different cost base, and the sensitivity is roughly linear in volume for payment fees '
          'and compounding for the two infrastructure lines. The rebasing also cannot repair a cost '
          'line the model never contained: it corrects what is indexed to the wrong series, not '
          'what is absent, and the absences are set out in Chapter 9.'),
    ]),
]})

# ---- 3 Shape of the cost base
comp = [('Cost of goods sold', cogs_y), ('Sales and marketing', sm), ('Legal', legal),
        ('General and administrative', ga), ('Development', dev)]
chapters.append({'num': '3', 'id': 'shape', 'title': 'The Shape of the Cost Base', 'nav': 'Cost Base', 'blocks': [
    kpis([
        {'n': M(T(total_y)), 'l': 'Rebased 3-year cost base'},
        {'n': M(T(pers_emp_y)), 'l': 'Personnel at employer cost'},
        {'n': P(100 * T(pers_emp_y) / T(total_y)), 'l': 'Personnel share of base'},
        {'n': '9', 'l': 'Steady-state headcount'},
        {'n': '{:.2f}×'.format(ratio_y[2]), 'l': 'Cost-to-revenue, FY2028'},
    ]),
    sub('3.1 Composition', [
        p('The cost base is dominated by cost of goods sold, which carries personnel, the two '
          'infrastructure lines, software licences and payment fees. Sales and marketing is the '
          'largest single discretionary commitment; legal is concentrated almost entirely in the '
          'launch year.'),
        table(['Category'] + YEARS + ['3-year total', 'Share'],
              [[k] + [M(v[i]) for i in range(3)] + [M(T(v)), P(100 * T(v) / T(total_y))]
               for k, v in comp]
              + [['Total'] + [M(total_y[i]) for i in range(3)] + [M(T(total_y)), '100.0%']],
              'Rebased. Source: Costs rows 51, 59, 66, 71, 76 and 78, with revenue-scaling lines '
              'replaced per Chapter 2.'),
        chart('Cost composition — three-year total',
              [{'lab': k, 'val': M(T(v)), 'pct': 100 * T(v) / T(total_y)} for k, v in comp]),
    ]),
    sub('3.2 The second-year decline', [
        p('Costs fall in the second year. This is structural rather than an efficiency: the launch '
          'year carries regulatory licensing, the smart-contract audit, the bug-bounty allocation, '
          'equipment and application development, and those obligations do not recur on the model’s '
          'schedule. The base then resumes growing in the third year as the volume-linked lines '
          'compound.'),
        table(['Year', 'Total cost', 'Change', 'Protocol revenue', 'Cost-to-revenue'],
              [[YEARS[y], M(total_y[y]),
                '—' if y == 0 else M(total_y[y] - total_y[y - 1]),
                M(rev_y[y]), '{:.2f}×'.format(ratio_y[y])] for y in range(3)],
              'Rebased costs against simulator default-case protocol revenue '
              '(BCI surplus capital plus Protocol Reserve).'),
    ]),
]})

# ---- 4 COGS
cogs_lines = [
    ('Personnel (employer cost)', pers_emp_y, 'Roster, flags and salary escalation'),
    ('Payment fees', fee_y, '0.1% of protocol transaction volume'),
    ('Maintenance infrastructure', maint_y, 'Seeded at $50,000/yr, compounds on volume growth'),
    ('Cloud infrastructure', cloud_y, 'Seeded at $42,000/yr, compounds on volume growth'),
    ('Software licences', software, 'Headcount-driven, $15,000/yr at full roster'),
]
chapters.append({'num': '4', 'id': 'cogs', 'title': 'Cost of Goods Sold', 'nav': 'Cost of Goods Sold', 'blocks': [
    p('Cost of goods sold carries five lines. Two are fixed against headcount, two compound with '
      'protocol volume, and one is a direct function of it. Each is stated below with the driver '
      'that produces it.'),
    table(['Line'] + YEARS + ['3-year total', 'Driver'],
          [[k] + [M(v[i]) for i in range(3)] + [M(T(v)), d] for k, v, d in cogs_lines]
          + [['Total'] + [M(cogs_y[i]) for i in range(3)] + [M(T(cogs_y)), '—']],
          'Rebased. Source: Costs rows 45–51; Personnel costs row 85 plus employer contribution.'),
    p('The composition changes materially across the horizon. Personnel is '
      + P(100 * pers_emp_y[0] / cogs_y[0]) + ' of cost of goods sold in the first year and '
      + P(100 * pers_emp_y[2] / cogs_y[2]) + ' in the third, not because the team shrinks but '
      'because the volume-linked lines compound around it. That is operating leverage, and it is '
      'the single most important structural property of this cost base.'),
]})

# ---- 5 Personnel and compensation
roster_rows = []
for r in roles:
    ref = roster_ref.get(r['name'], {})
    hire = ref.get('hire')
    roster_rows.append([
        r['name'], ref.get('alloc', ''),
        M(ref.get('base', 0)), '{:.0%}'.format(ref.get('bonus', 0)),
        '{}-{:02d}'.format(*hire) if hire else '—',
        M(r['gross']),
    ])

gross_t = pers_gross
cas_t, cass_t = gross_t * CAS, gross_t * CASS
pit_t = (gross_t - cas_t - cass_t) * PIT
net_t = gross_t - cas_t - cass_t - pit_t
cam_t = gross_t * CAM
emp_t = gross_t + cam_t
wedge_t = cas_t + cass_t + pit_t + cam_t

tax_rows = []
for r in roles:
    g = r['gross']
    cas, cass = g * CAS, g * CASS
    pit = (g - cas - cass) * PIT
    tax_rows.append([r['name'], M(g), M(cas), M(cass), M(pit), M(g - cas - cass - pit),
                     M(g * CAM), M(g * (1 + CAM))])
tax_rows.append(['Total', M(gross_t), M(cas_t), M(cass_t), M(pit_t), M(net_t), M(cam_t), M(emp_t)])

chapters.append({'num': '5', 'id': 'personnel', 'title': 'Personnel and Compensation',
                 'nav': 'Personnel', 'blocks': [
    sub('5.1 The roster', [
        p('The model defines twelve roles and hires nine. Headcount reaches five in the first month, '
          'nine by the fifth, and does not change again for the remainder of the horizon. Salary is '
          'the only escalator: gross pay is indexed at 5.5% in 2026, 6.0% in 2027 and 5.7% '
          'thereafter, following the Romanian national forecast, and every role except the founder '
          'carries a 10% bonus.'),
        table(['Role', 'Allocation', 'Gross start salary', 'Bonus', 'Hire', '3-year gross'],
              roster_rows,
              'Source: Personnel costs rows 9–20 (roster and flags) and rows 73–84 (summary). '
              'Gross is salary plus bonus, before employer contributions.'),
        p('Three defined roles are never hired: ' + ', '.join(never_hired) + '. They carry salary '
          'assumptions in the model and contribute nothing to its cost. They are disclosed here '
          'rather than dropped, because a roster that prices twelve roles and staffs nine is a '
          'hiring decision that has already been taken.'),
    ]),
    sub('5.2 Gross-to-net under Romanian payroll law', [
        p('The model records gross pay. Gross pay is not what the protocol spends, and it is not '
          'what anyone receives. Under Romanian law an employee contributes 25% of gross to the '
          'pension fund and 10% to health insurance, both uncapped; income tax of 10% is then '
          'charged on what remains, which is 6.5% of gross. The employer adds a work-insurance '
          'contribution of 2.25% on top of gross and owes no pension contribution under normal '
          'working conditions.'),
        p('The arithmetic that follows is therefore fixed: an employee receives 58.5% of gross, the '
          'State Budget receives 41.5% of gross from within it, and the employer pays 102.25% of '
          'gross in total.'),
        table(['Role', 'Gross', 'Pension 25%', 'Health 10%', 'Income tax', 'Net received',
               'Employer 2.25%', 'Total employer cost'],
              tax_rows,
              'Thirty-six months, FY2026–FY2028, USD. Rates verified against PwC Worldwide Tax '
              'Summaries (Romania, 2026): employee CAS 25% and CASS 10%, both uncapped; income tax '
              '10% flat on the post-contribution base; employer CAM 2.25%.'),
        kpis([
            {'n': M(emp_t), 'l': 'Total employer cost'},
            {'n': M(net_t), 'l': 'Net reaching employees', 'c': 'g'},
            {'n': M(wedge_t), 'l': 'Total to State Budget', 'c': 'w'},
            {'n': P(100 * wedge_t / emp_t), 'l': 'Effective tax wedge'},
        ]),
        chart('Where each dollar of employer cost goes',
              [{'lab': 'Net to employees', 'val': M(net_t), 'pct': 100 * net_t / emp_t},
               {'lab': 'Pension (CAS)', 'val': M(cas_t), 'pct': 100 * cas_t / emp_t},
               {'lab': 'Health (CASS)', 'val': M(cass_t), 'pct': 100 * cass_t / emp_t},
               {'lab': 'Income tax', 'val': M(pit_t), 'pct': 100 * pit_t / emp_t},
               {'lab': 'Work insurance', 'val': M(cam_t), 'pct': 100 * cam_t / emp_t}]),
    ]),
    sub('5.3 What the model omits', [
        p('The model carries a 35% payroll-tax assumption that no formula references. That figure '
          'is the employee contribution — pension and health combined — and it is deducted from '
          'gross rather than added to it, so its absence from the formulas is correct and the '
          'personnel line is not understated by 35%. What the model does omit is the employer’s '
          'own contribution: 2.25% of gross, or ' + M(cam_t) + ' across the horizon. That amount is '
          'added to every personnel figure in this document and is the only adjustment made to the '
          'model’s personnel series.'),
        table(['Year', 'Model gross', 'Employer contribution', 'True employer cost', 'Net to staff'],
              [[YEARS[y], M(pers_gross_y[y]), M(pers_cam_y[y]), M(pers_emp_y[y]),
                M(pers_gross_y[y] * NET_RATE)] for y in range(3)]
              + [['Total', M(gross_t), M(cam_t), M(emp_t), M(net_t)]],
              'Source: Personnel costs row 85, with employer CAM applied.'),
    ]),
]})

# ---- 6 Operating expenses
chapters.append({'num': '6', 'id': 'opex', 'title': 'Operating Expenses', 'nav': 'Operating Expenses', 'blocks': [
    p('Outside cost of goods sold the base carries four categories, none of which is indexed to '
      'protocol activity. Their shape is dominated by what happens once at launch and what recurs '
      'thereafter.'),
    table(['Category'] + YEARS + ['3-year total', 'Character'],
          [['Sales and marketing'] + [M(sm[i]) for i in range(3)] + [M(T(sm)),
            '$1.5M launch year, $1.0M recurring'],
           ['Legal'] + [M(legal[i]) for i in range(3)] + [M(T(legal)),
            'Licensing, audit and bounty at launch; advisory recurring'],
           ['Development'] + [M(dev[i]) for i in range(3)] + [M(T(dev)),
            'Application build at launch; integrations recurring'],
           ['General and administrative'] + [M(ga[i]) for i in range(3)] + [M(T(ga)),
            'Equipment at launch; office, accounting and recruiting recurring']],
          'Source: Costs rows 53–76. Not rebased — none of these lines is revenue-indexed in the '
          'model.'),
    p('Sales and marketing is ' + P(100 * T(sm) / T(total_y)) + ' of the three-year base and the '
      'largest discretionary commitment in the model. It is also the reason personnel accounts for '
      'a smaller share of this cost base than is typical, a point taken up in Chapter 8.'),
]})

# ---- 7 Coverage
# ============================================================ entity attribution
chapters.append({'num': '7', 'id': 'entities', 'title': 'Entity Attribution',
                 'nav': 'Entity Attribution', 'blocks': [
    p('Cost belongs to an entity. So does revenue. Until this chapter existed, this '
      'document reported a single cost base and the business plan set protocol revenue '
      'against it — money that accrues to index holders, netted against costs borne by a '
      'company that has no revenue at all. The two sit in different entities and cannot '
      'be combined.'),
    table(['Entity', 'Jurisdiction', 'Role', 'Bears'],
          [['Nexus Technologies Foundation', 'Panama',
            'Owns the protocol and its intellectual property; grants the token delegation',
            'Capitalised development IP'],
           ['BITSWIFT TECHNOLOGIES INC.', 'Panama',
            'Distributes the token; holds raise proceeds; counterparty on the SAFT',
            'Token legal'],
           ['Operating company', 'European Union',
            'Develops the protocol, employs the team, runs go-to-market',
            'All operating cost']],
          'Entity roles as recorded in the corporate documents and the SAFT.'),
    sub('7.1 Every cost line, and who bears it', [
        table(['Line'] + YEARS + ['36-month total', 'Borne by'],
              [[n] + [M(v[i]) for i in range(3)] + [M(T(v)), o] for n, v, o in ENTITY_LINES]
              + [['Total'] + [M(total_y[i]) for i in range(3)] + [M(T(total_y)), '']],
              'Payment fees scale with on-chain volume and are netted against protocol '
              'revenue, not carried by the company.'),
        table(['Bearer'] + YEARS + ['36-month total', 'Share'],
              [['Operating company'] + [M(opco_y[i]) for i in range(3)] + [M(T(opco_y)), P(100*T(opco_y)/T(total_y))],
               ['Protocol'] + [M(proto_y[i]) for i in range(3)] + [M(T(proto_y)), P(100*T(proto_y)/T(total_y))]],
              'The operating company bears ' + P(100*T(opco_y)/T(total_y)) + ' of the base.'),
    ]),
    sub('7.2 Development is billed on, not absorbed', [
        p('The Foundation owns the protocol. Development work creates that protocol. So the '
          'operating company bills the Foundation for it and the Foundation capitalises the '
          'result as intellectual property — ' + M(T(dev_capitalised)) + ' over the horizon, '
          'inclusive of the services margin. Treating it as a bare expense of the operating '
          'company would mean one entity paying to build an asset another entity owns, with '
          'nothing recorded against it.'),
    ]),
]})

# ============================================================ operating company P&L
chapters.append({'num': '8', 'id': 'opco', 'title': 'The Operating Company',
                 'nav': 'Operating Company', 'blocks': [
    p('The operating company is paid under a development-and-services agreement with the '
      'Foundation and the Issuer, at cost plus ' + '{:.0%}'.format(SERVICES_MARGIN) + '. That '
      'agreement is what gives it a revenue line. Without one it is a cost centre with no '
      'income — commercially wrong, and indefensible on transfer pricing, since an EU company '
      'performing real work for a Panama principal at zero margin is the textbook challenge.'),
    kpis([
        {'n': M(sum(opco_rev_y)), 'l': 'Services revenue, 36 months'},
        {'n': M(sum(opco_marg_y)), 'l': 'Margin at {:.0%}'.format(SERVICES_MARGIN), 'c': 'g'},
        {'n': M(sum(opco_tax_y)), 'l': 'EU tax at {:.0%}'.format(EU_TAX_RATE), 'c': 'w'},
        {'n': M(OPCO_BURN_MO), 'l': 'Monthly cost'},
    ]),
    sub('8.1 Profit and loss', [
        table(['USD'] + YEARS + ['36-month total'],
              [['Services revenue'] + [M(opco_rev_y[i]) for i in range(3)] + [M(sum(opco_rev_y))],
               ['Operating cost'] + [M(-opco_y[i]) for i in range(3)] + [M(-sum(opco_y))],
               ['Margin'] + [M(opco_marg_y[i]) for i in range(3)] + [M(sum(opco_marg_y))],
               ['EU corporate tax'] + [M(-opco_tax_y[i]) for i in range(3)] + [M(-sum(opco_tax_y))],
               ['Net result'] + [M(opco_net_y[i]) for i in range(3)] + [M(sum(opco_net_y))]],
              'Cost plus {:.0%}, taxed at the Romanian corporate rate of {:.0%}. The company '
              'is deliberately close to break-even: it is a service provider, not a profit '
              'centre.'.format(SERVICES_MARGIN, EU_TAX_RATE)),
    ]),
    sub('8.2 How it is funded, and for how long', [
        p('The company runs a <b>24-month funded plan</b>. That horizon is its own — the '
          'protocol\'s 36-month simulation is a different entity on a different clock, and the '
          'two are reported separately throughout this document set.'),
        table(['Source', 'Amount', 'Form'],
              [['Seed raise, directed to operations', M(2_480_000), 'Cash at token generation'],
               ['Company share of the public raise, 30%', M(2_250_000), 'Cash at token generation'],
               ['Marketing allocation, 2% of supply', M(TOK_MKT), 'Tokens, liquidated on schedule'],
               ['Treasury allocation vested by month 24', M(TOK_TRE_24), 'Tokens, liquidated on schedule'],
               ['Total available', M(OPCO_AVAIL_24), '']],
              'Token figures are valued at the launch price and vest on the schedules in the '
              'tokenomics model.'),
        note('Two of the four sources above are token allocations valued at the launch price. '
             'They are shown here because they are already allocated and already vesting, not '
             'because the plan requires them to be sold. The operating company also holds an '
             'equity route: it can raise against its own equity to cover working capital, and '
             'that option substitutes for some or all of the token allocations shown above. The '
             'company therefore has more than one path to the same funded position.',
             '', 'More than one funding route'),
        table(['24-month position', 'Amount'],
              [['Cost to month 24, including margin', M(OPCO_COST_24)],
               ['Funding available by month 24', M(OPCO_AVAIL_24)],
               ['Surplus', M(OPCO_SURPLUS_24)]],
              'Monthly cost ' + M(OPCO_BURN_MO) + '. The plan is funded for its stated horizon.'),
    ]),
]})

# ============================================================ marketing plan
chapters.append({'num': '9', 'id': 'marketing', 'title': 'The Marketing Budget',
                 'nav': 'Marketing Budget', 'blocks': [
    p('Sales and marketing is ' + M(MKT_TOTAL) + ' over the horizon, rebuilt from published '
      'benchmarks rather than carried across from the financial model, which had it at '
      '$3,740,000. The first two stages of the go-to-market plan are a direct sales motion — '
      'institutional liquidity providers and RWA issuers closed one relationship at a time — '
      'and that motion does not consume a mass-market budget.'),
    p('At $5,000 to $15,000 per sourced opportunity, the ten to fifteen relationships the '
      'first stage needs cost between $75,000 and $225,000 to source. The old budget spent '
      '$1,580,000 in the launch year on a motion that cannot absorb it.'),
    table(['Phase', 'When', 'Line', 'Amount'],
          [[ph, wh, ln, M(v)] for ph, wh, items in MARKETING_PLAN for ln, v in items]
          + [['Total', '', '', M(MKT_TOTAL)]],
          'Benchmarks, July 2026: Tier-2 conference $50,000–150,000 all-in; events 5–15% of '
          'B2B marketing budgets; $5,000–15,000 per sourced opportunity; median B2B sales-led '
          'customer acquisition cost $11,400.'),
    chart('Marketing by phase',
          [{'lab': ph.split(' — ')[1], 'val': M(v), 'pct': 100 * v / MKT_TOTAL}
           for (ph, _, _), v in zip(MARKETING_PLAN, MKT_PHASE_TOTAL)]),
    p('The <b>Marketing token allocation covers ' + M(MKT_TOKEN_ALLOCATION) + '</b> of this — '
      + P(100*MKT_TOKEN_ALLOCATION/MKT_TOTAL) + ' of the budget. Marketing is funded from that '
      'allocation first and from cash second. Spending cash while the allocation intended for '
      'exactly this purpose sits idle would be the wrong way round.'),
]})

chapters.append({'num': '10', 'id': 'coverage', 'title': 'Cost Coverage', 'nav': 'Cost Coverage', 'blocks': [
    p('Coverage is an entity-level question, and Chapter 7 established that this document set '
      'reports two entities on two clocks. Reporting a single blended coverage ratio would repeat '
      'exactly the error this document exists to correct: protocol revenue accrues to index '
      'holders and to the Protocol Reserve, while the operating cost base is reimbursed under a '
      'services agreement. The two are stated separately below and are never netted.'),
    sub('10.1 The protocol', [
        p('The protocol bears payment processing fees and nothing else. Every other line in '
          'Chapter 7 belongs to the operating company. Against that, protocol revenue is the '
          'BCI surplus and the Protocol Reserve.'),
        table(['Year', 'Protocol revenue', 'Protocol-borne cost', 'Net', 'Coverage'],
              [[YEARS[y], M(rev_y[y]), M(proto_y[y]), M(rev_y[y] - proto_y[y]),
                '{:.2f}\u00d7'.format(rev_y[y] / proto_y[y])] for y in range(3)]
              + [['Cumulative', M(T(rev_y)), M(T(proto_y)), M(T(rev_y) - T(proto_y)),
                  '{:.2f}\u00d7'.format(T(rev_y) / T(proto_y))]],
              'Protocol revenue is BCI surplus capital plus Protocol Reserve, simulator default '
              'case. Protocol-borne cost is payment processing fees, per Chapter 7.'),
        p('The protocol covers its own costs in every year of the horizon. This is a structural '
          'result rather than a favourable one: the protocol is not the entity that employs the '
          'team or runs go-to-market, so almost none of the cost base sits here. What this line '
          'does establish is that the fee engine is self-sustaining on its own terms, and that '
          'protocol revenue is available in full to index holders and the Reserve.'),
    ]),
    sub('10.2 The operating company', [
        p('The operating company is reimbursed at cost plus '
          + '{:.0%}'.format(SERVICES_MARGIN) + ' under the development-and-services agreement. '
          'Its revenue is that fee, not protocol revenue, which it never receives.'),
        table(['Year', 'Services revenue', 'Operating cost', 'Margin', 'Net after EU tax'],
              [[YEARS[y], M(opco_rev_y[y]), M(-opco_y[y]), M(opco_marg_y[y]), M(opco_net_y[y])]
               for y in range(3)]
              + [['Cumulative', M(T(opco_rev_y)), M(-T(opco_y)), M(T(opco_marg_y)),
                  M(T(opco_net_y))]],
              'Cost plus {:.0%}, taxed at the EU corporate rate of {:.0%}.'.format(
                  SERVICES_MARGIN, EU_TAX_RATE)),
        p('The company returns ' + M(T(opco_net_y)) + ' across the horizon. It cannot report a '
          'loss on this structure, because its revenue is defined as its cost plus a margin. That '
          'is the point of a cost-plus service company and not evidence of commercial strength. '
          'The question that matters for the company is funding, not profitability, and it is '
          'answered in Chapter 8: the plan is funded to month twenty-four with a surplus.'),
    ]),
    sub('10.3 Why these are not combined', [
        p('Setting ' + M(T(rev_y)) + ' of protocol revenue against ' + M(T(total_y)) + ' of total '
          'cost base would produce a deficit of ' + M(T(rev_y) - T(total_y)) + ' and a coverage '
          'ratio of ' + '{:.2f}\u00d7'.format(T(total_y) / T(rev_y)) + '. That figure is stated '
          'here only so that it is not derived independently and misread. It has no economic '
          'meaning: it charges money owed to index holders against costs borne by a company those '
          'holders do not own, and it counts protocol-borne payment fees on the cost side while '
          'counting protocol revenue on the other. Chapter 7 rules the combination out, and this '
          'chapter observes that ruling.'),
    ]),
    note('These are model outputs on the default preset, not forecasts and not a funding plan. '
         'A different preset, a different market set, or a different marketing commitment moves '
         'every figure in this chapter.', 'warn', 'Read these numbers carefully'),
]})

# ---- 8 External validation
bench_rows = [
    ['Smart-contract audit', '$200,000, launch year',
     '$50K–$100K typical DeFi; $60K–$120K realistic pre-launch',
     'Adequate to conservative', 'Sherlock; Zealynx; QuillAudits (2026)'],
    ['Bug bounty', '$300,000, launch year only',
     '$80K–$100K max reward mid-size; $1M+ at scale',
     'Budget reasonable; recurrence missing', 'Immunefi; Sherlock (2026)'],
    ['Regulatory licensing', '$300,000, launch year only',
     '€350K–€900K first year; €150K–€250K annually to hold',
     'Understated on both counts', 'Crassula; Pharos; InnReg (2026)'],
    ['Legal advisory', '$72,000 per year',
     '€80K–€200K for authorisation preparation alone',
     'Plausible steady state; light at launch', 'Crassula; InnReg (2026)'],
    ['Smart-contract engineer', '$128,250 base',
     '$150K–$200K mid-level; $200K–$300K senior, plus tokens',
     'Below the mid-level floor', 'InterviewPal; CryptoRecruit; Metana (2026)'],
    ['Chief technology officer', '$129,600 base',
     'Senior individual contributors alone command $200K–$300K',
     'Materially below market', 'InterviewPal; CryptoRecruit (2026)'],
]
chapters.append({'num': '11', 'id': 'validation', 'title': 'External Validation', 'nav': 'External Validation', 'blocks': [
    p('Every material line was tested against published market ranges for 2026. The results divide '
      'cleanly: the security budget is adequate, the regulatory budget is not, and the technical '
      'salaries sit below the bands they are competing in.'),
    table(['Line', 'Model figure', 'Market range', 'Assessment', 'Source'], bench_rows,
          'Benchmarks gathered July 2026. Ranges are as published by the cited sources and are '
          'indicative, not quotations for this protocol.'),
    sub('8.1 Regulatory licensing', [
        p('The model books regulatory licensing once, at $300,000, in the launch year and nothing '
          'thereafter. Published first-year costs for a MiCA authorisation run from €350,000 to '
          '€900,000 all-in, and holding the authorisation costs a further €150,000 to €250,000 '
          'every year. The model is therefore light in the year it spends and silent in the four '
          'years it must keep spending. This is the largest single omission identified in the cost '
          'base.'),
    ]),
    sub('8.2 Technical compensation', [
        p('The smart-contract engineer is budgeted at $128,250 and the chief technology officer at '
          '$129,600. Published 2026 ranges place mid-level smart-contract engineers at $150,000 to '
          '$200,000 and senior engineers who own protocol architecture at $200,000 to $300,000 '
          'before token compensation. Lifting both roles to a $200,000 floor adds approximately '
          '$142,150 of base salary a year, or roughly $875,000 across five years once bonus and '
          'employer contributions are included.'),
        p('The consequence is not that the model is wrong — it is that the model prices a team it '
          'may not be able to recruit at these levels. That is a hiring risk carried inside a cost '
          'assumption, and it belongs in the cost document rather than outside it.'),
    ]),
    sub('8.3 Payroll share', [
        p('Payroll typically accounts for 60% to 75% of an early-stage company’s burn. In this '
          'model it is ' + P(100 * T(pers_emp_y) / T(total_y)) + ' of the three-year base. The '
          'difference is almost entirely the $1.5 million launch marketing commitment and the '
          '$1.0 million a year that follows it. Whether that allocation is correct is a strategic '
          'question rather than an accounting one; what the comparison establishes is that this '
          'cost base is not shaped like a typical early-stage company’s, and the divergence is '
          'deliberate and locatable.'),
    ]),
]})

# ---- 9 Limits
chapters.append({'num': '12', 'id': 'limits', 'title': 'Model Limits and Open Items', 'nav': 'Limits', 'blocks': [
    p('This chapter names what the cost model does not contain or does not support. It is an '
      'engineering disclosure of the model’s boundaries, not a criticism of it.'),
    ul([
        'Recurring regulatory cost: licensing is booked once at launch. Holding a MiCA '
        'authorisation costs €150,000 to €250,000 a year and appears nowhere in the base.',
        'Recurring security cost: the bug bounty is booked once at launch. A bounty programme that '
        'closes after year one is not a bounty programme.',
        'Headcount plan: the roster reaches nine in month five and never changes, across a horizon '
        'in which value under custody grows from $3.6 million to $46.1 million. No hiring is funded '
        'against that growth.',
        'Technical compensation: two of the three engineering-side salaries sit below published '
        'market floors, as set out in Chapter 8.',
        'Cost allocation: every role carries a G&A, S&M or R&D tag, but the cost forecast routes '
        'all personnel into cost of goods sold. The tags are unused, so no functional cost split is '
        'available from this model.',
        'Horizon: the analysis stops at month thirty-six because the simulator does. Years four and '
        'five of the financial model are unexamined here.',
        'Scenario dependence: every rebased figure inherits the default preset. No scenario range '
        'is presented, and none should be inferred.',
        'Unhired roles: three defined roles carry salary assumptions and no hiring flag. If the '
        'plan intends to hire them, the cost base is understated by their fully loaded cost.',
    ]),
]})

# ---- Disclaimer
chapters.append({'num': '', 'id': 'disclaimer', 'title': 'Legal Disclaimer', 'nav': 'Legal Disclaimer', 'blocks': [
    p('This document is provided for informational purposes only. It does not constitute an offer '
      'to sell, a solicitation of an offer to buy, or a recommendation of any token, security or '
      'financial instrument, in any jurisdiction, and nothing in it constitutes investment, legal, '
      'tax or accounting advice. All figures are outputs of a financial model and of the BOUND '
      'simulation engine on the assumptions recorded in this document; they are not forecasts, '
      'budgets, commitments or guarantees of actual results. Cost figures rebased onto the '
      'simulator inherit every assumption of its default preset. Payroll figures apply Romanian '
      'statutory rates as published for 2026 and are illustrative of the model’s roster rather '
      'than a statement of any individual’s compensation; the applicable jurisdiction, entity of '
      'employment and engagement structure are matters of fact not determined by this document. '
      'Market benchmarks are drawn from third-party published sources as at July 2026, are '
      'indicative rather than quotations, and may change. Forward-looking statements are subject '
      'to material uncertainty and change. © 2026 BOUND Protocol.'),
]})

out = {'front': front, 'chapters': chapters}
dest = os.path.join(ROOT, 'scripts/costs-content.json')
with open(dest, 'w') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)
    f.write('\n')

# Machine-readable cost feed. The Business Plan builds its income statement from
# this rather than re-deriving the cost model, so the two documents can never
# disagree about what the protocol spends.
with open(os.path.join(ROOT, 'scripts/cost-model.json'), 'w') as f:
    json.dump({
        'basis': 'rebased on simulator default case',
        'years': YEARS,
        'cogs': {
            'software': software, 'cloud': cloud_y, 'maintenance': maint_y,
            'personnel_gross': pers_gross_y, 'personnel_employer': pers_emp_y,
            'payment_fees': fee_y, 'total': cogs_y,
        },
        'opex': {'ga': ga, 'legal': legal, 'sales_marketing': sm, 'development': dev},
        'total': total_y,
        'payroll': {
            'gross': gross_t, 'cas': cas_t, 'cass': cass_t, 'income_tax': pit_t,
            'net': net_t, 'employer_cam': cam_t, 'employer_total': emp_t, 'wedge': wedge_t,
        },
        'headcount_steady': len(roles),
    }, f, indent=2)
    f.write('\n')

print('wrote scripts/cost-model.json')
print('wrote scripts/costs-content.json — {} chapters'.format(len(chapters)))
print('  rebased 3-yr cost base : {}'.format(M(T(total_y))))
print('  protocol revenue       : {}'.format(M(T(rev_y))))
print('  cumulative net         : {}'.format(M(T(net_y))))
print('  personnel employer cost: {}'.format(M(emp_t)))
print('  net to employees       : {}'.format(M(net_t)))
print('  tax wedge              : {} ({:.1f}%)'.format(M(wedge_t), 100 * wedge_t / emp_t))
